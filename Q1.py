import os
import re
from nltk.stem import WordNetLemmatizer
import time
import numpy as np
import createindex
import math
import ast
from openpyxl import Workbook,load_workbook
import win32com.client
from scipy.spatial import distance
from tkinter import Scale, Tk, Frame, Label, Button
from tkinter import *
from tkinter.ttk import Notebook,Entry

def programRun(valk,tr,tes):        #function to run the whole program if a training or testing ratio is changed
    createindex.splitratio(tr,tes)      #splits the bbc folder into train and test folders
    createindex.listofwords()           #creste the vocabulary fille based on the training data 
    lemmatizing=WordNetLemmatizer ()

    def returndict():                       #return the vocabulary of the training data
        f=open("vocabulary.txt","r")        #open file
        r=f.read()                            #read file
        dc=ast.literal_eval(r)                  #to get the dictionary from the txt file in the correct form 
        f.close()                           #close file
        return dc
    #dictionary of the terms
    Dictionary=returndict()


    def returnfilelist():                      #returns the list of files present in the training data
        filee=open("filelist.txt","r")          #open file
        r=filee.read()                          #read file
        r=ast.literal_eval(r)                  #to get the file list from the txt file in the correct form
        filee.close()                       # file close
        return r                            

    filelist=returnfilelist()

    #returns a list of words in sorted order so that we can check the index of a particular word

    def returnlistofwords():                        #return a list of words in the sorted order , so that we can get the correct position of the term in vector
        fl=open("listfwords.txt","r")
        r=fl.read()
        r=ast.literal_eval(r)
        fl.close()
        return r

    lstofword=returnlistofwords()

    def get_df():                                      #returns a list of document frequency of terms
        templist=[]
        for i in range(len(lstofword) ):
            templist.append(0)
        for i in lstofword:
            templist[lstofword.index(i)]=( len(Dictionary[i]))
        return templist

    dflist=get_df()         #list of doc frequency of terms

    N=len(filelist)        #total number of documents 

    def VSMofDoc():
        dict={}                      #dictionary for VSM
        temp=[]                     #temp list for document vector
        path = os.getcwd()          #returns the path of the current folder 
        path+="\\output\\train"
        folders=os.listdir(path)
        qq=re.compile("[^a-zA-Z]")

        for folder in folders:              # to iterate through folder
            fname=path+"\\"+folder
            allfiles=os.listdir(fname)      #all file present in the current folder

            for f in allfiles:                     # for each file present in the directory
                for jj in range(len(lstofword)):    # initilizing the vector with zero 
                    temp.append(0)
                
                fname=path+"\\"+folder
                fname+="\\"+f                            #file name
                lines=open(fname,"r")               #open a file 
                wordslist=lines.read()              #read the data from the file
                line = wordslist.rstrip()
                # x = re.findall('[a-zA-Z]*', line)
                x=qq.split(line)
                x = list(filter(None, x))          

                for i in x:
                    i=i.lower()
                    i=lemmatizing.lemmatize(i)      #lemmatizing the word
                    if i in lstofword:
                        indexofterm=lstofword.index(i)      #getting the index of the term
                        temp[indexofterm]+=1            #incremening the term frequency
                name=re.findall('[0-9]*',f)
                dict[folder+"\\"+name[0]]=temp.copy()                 #assigning the vector to its doc id
                lines.close()                       #closing the file
                fname=""
                temp.clear()
        
        # for ii,j in dict.items():
        #     for iii in range(len(j)):
        #         if j[iii]!=0:
        #             product=( j[iii] )*( math.log( N/dflist[iii] ,2 ) )           #mutiplying the term frequency by its idf
        #             j[iii]=product
        
        #write the VSM to a file  
        wordfile=open("DocumentVector.txt","w")
        wordfile.write(str(dict))
        wordfile.close()
        return dict
        # return
    dictofTrain=VSMofDoc()

    #read the VSM file
    # f=open("DocumentVector.txt","r")
    # r=f.read()
    # dictofTrain=ast.literal_eval(r)
    # f.close()

    def VSMofTest():
        dict={}                      #dictionary for VSM
        temp=[]                     #temp list for document vector
        global Testing_filelist
        Testing_filelist=[]
        path = os.getcwd()          #returns the path of the current folder 
        path+="\\output\\val"
        folders=os.listdir(path)
        
        for folder in folders:
            fname=path+"\\"+folder
            allfiles=os.listdir(fname)

            for f in allfiles:                     # for each file present in the directory
                for jj in range(len(lstofword)):    # initilizing the vector with zero 
                    temp.append(0)
                
                fname=path+"\\"+folder
                fname+="\\"+f                            #file name
                lines=open(fname,"r")               #open a file 
                wordslist=lines.read()              #read the data from the file
                line = wordslist.rstrip()
                x = re.findall('[a-zA-Z]*', line)
                x = list(filter(None, x))  
                parsedquery=[]        
                for i in x:
                    i=i.lower()
                    i=lemmatizing.lemmatize(i.lower())      #lemmatizing the word
                    if i in lstofword:
                        indexofterm=lstofword.index(i)      #getting the index of the term
                        temp[indexofterm]+=1            #incremening the term frequency

                name=re.findall('[0-9]*',f)
                dict[folder+"\\"+name[0]]=temp.copy()                 #assigning the vector to its doc id
                Testing_filelist.append(folder+"\\"+name[0])
                lines.close()                       #closing the file
                fname=""
                temp.clear()
        
        for ii,j in dict.items():
            for iii in range(len(j)):
                if j[iii]!=0:
                    product=( j[iii] )*( math.log( N/dflist[iii] ,2 ) )           #mutiplying the term frequency by its idf
                    j[iii]=product
        
        #write the VSM to a file  
        wordfile=open("TestingDocVector.txt","w")
        wordfile.write(str(dict))
        wordfile.close()
        return dict
        # return
    dictofTest=VSMofTest()

    #read the VSM file
    # f=open("TestingDocVector.txt","r")
    # r=f.read()
    # dictofTest=ast.literal_eval(r)
    # f.close()
    # print(len(Testing_filelist))
    # print(len(filelist))

    def euclidean_distance(K):
        WB_Ans=Workbook()
        for i in Testing_filelist:
            minlist=[[np.inf for i in range(2)] for j in range(K)]
            sheetname=i.replace("\\", " ")
            sheet=WB_Ans.create_sheet(sheetname)
            sheet.cell(row=1,column=1).value="Testing Doc"
            sheet.cell(row=1,column=2).value="Training Doc"
            sheet.cell(row=1,column=3).value="Euclidean Distance"
            sheet.cell(row=1,column=4).value="K-Nearest Neighbour"
            sheet.cell(row=1,column=5).value="Actual"
            sheet.cell(row=1,column=6).value="Predicted"
            row=2
            for j in filelist:
                d=distance.euclidean(dictofTrain[j], dictofTest[i])
                sheet.cell(row=row,column=1).value=i
                sheet.cell(row=row,column=2).value=j
                sheet.cell(row=row,column=3).value=d
                maxval=max(minlist, key=lambda x: x[1])
                if maxval[1]>d:
                    fn=j.split("\\")
                    minlist[minlist.index(maxval)]=[fn[0],d]
                row+=1
                
            minlist.sort(key = lambda x: x[1])
            row=2
            for i in minlist:
                sheet.cell(row=row, column=4).value=i[0]+" ("+str(i[1])+")"
                row+=1
                
        WB_Ans.save(filename="Answerfilenew.xlsx")
        
    K=valk
    euclidean_distance(K)
    def sortingAnswerfile(): 
        path=os.getcwd()
        path+="\\Answerfilenew.xlsx"
        excel = win32com.client.Dispatch("Excel.Application")
        wb = excel.Workbooks.Open(path)

        for i in Testing_filelist:
            sheetname=i.replace("\\", " ")
            ws = wb.Worksheets(sheetname)
            ws.Range("A:C").Sort(Key1=ws.Range("C1"),Order1=1,Header=1, Orientation=2)
        wb.Save()
        excel.Application.Quit()

    sortingAnswerfile()
    path=os.getcwd()
    path+="\\Answerfilenew.xlsx"
    WB=load_workbook(path)
    answer3.delete(1.0,END)
    confusionMatrix = [[0 for i in range(6)] for j in range(6)] 

    for i in Testing_filelist:
        sheetname=i.replace("\\", " ")
        ws=WB[sheetname]
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 25
        ws.column_dimensions["E"].width = 18
        ws.column_dimensions["F"].width = 18
        actual=sheetname.split()
        ws.cell(row=2,column=5).value=actual[0]
        
        Answerfrequency={}
        rr=2
        for i in range(K):   
            val=ws.cell(row=rr,column=4).value
            f=val.split(" ")
            if f[0] not in Answerfrequency:
                Answerfrequency[f[0]]=1
            else:
                Answerfrequency[f[0]]+=1
            rr+=1
        
        maxfrequency=1
        key=""
        for k, v in Answerfrequency.items():
            if v>maxfrequency:
                maxfrequency=v
                key=k
        if key=="":
            val=ws.cell(row=2,column=4).value
            f=val.split(" ")
            key=f[0]
        
        ws.cell(row=2,column=6).value= key
        Answerfrequency.clear()   
        if ws.cell(row=2,column=5).value=="athletics":
            r=0
        elif ws.cell(row=2,column=5).value=="cricket":
            r=1
        elif ws.cell(row=2,column=5).value=="football":
            r=2
        elif ws.cell(row=2,column=5).value=="rugby":
            r=3
        elif ws.cell(row=2,column=5).value=="tennis":
            r=4
            
        if ws.cell(row=2,column=6).value=="athletics":
            c=0
        elif ws.cell(row=2,column=6).value=="cricket":
            c=1
        elif ws.cell(row=2,column=6).value=="football":
            c=2
        elif ws.cell(row=2,column=6).value=="rugby":
            c=3
        elif ws.cell(row=2,column=6).value=="tennis":
            c=4
        answer3.insert(END,ws.cell(row=2,column=1).value)
        answer3.insert(END,'       ---       ')
        answer3.insert(END,ws.cell(row=2,column=6).value)
        answer3.insert(END,'\n')
        
        confusionMatrix[r][c]+=1
        confusionMatrix[5][5]=0
        for x in range(5):
            confusionMatrix[5][x]=0
            confusionMatrix[x][5]=0
            for y in range(5):
                confusionMatrix[5][x]+=confusionMatrix[y][x]
                confusionMatrix[x][5]+=confusionMatrix[x][y]
            
            confusionMatrix[5][5]+=confusionMatrix[5][x]
    WB.save("Answerfilenew.xlsx")
    # print(confusionMatrix)
    answer3.pack()
    
    Accuracy=0
    for i in range(len(confusionMatrix)-1):
        Accuracy+=confusionMatrix[i][i]

    Accuracy=Accuracy/(confusionMatrix[5][5])
    # print(Accuracy)
    answer4.delete(1.0,END)
    answer4.insert(END,'Rows = Actual ')
    answer4.insert(END,'\n')
    answer4.insert(END,'Columns = Predicted ')
    answer4.insert(END,'\n')
    answer4.insert(END,'\n')
    answer4.insert(END,'              ')
    answer4.insert(END,"Athletics   ")
    answer4.insert(END,"Cricket     ")
    answer4.insert(END,"Football    ")
    answer4.insert(END,"Rugby     ")
    answer4.insert(END,"Tennis    ")
    answer4.insert(END,"Total     ")
    answer4.insert(END,'\n')
    for i in range(len(confusionMatrix)):
        if i ==0:
            answer4.insert(END,"Athletics    ")
        elif i==1:
            answer4.insert(END,"Cricket       ")
        elif i==2:
            answer4.insert(END,"Football      ")
        elif i==3:
            answer4.insert(END,"Rugby         ")
        elif i==4:
            answer4.insert(END,"Tennis        ")
        elif i==5:
            answer4.insert(END,"Total         ")
        for j in range(len(confusionMatrix)):
            answer4.insert(END,confusionMatrix[i][j])
            answer4.insert(END,'             ')
        answer4.insert(END,'\n')
        
    answer4.insert(END,'\n')
    answer4.insert(END,'Accuracy  : ')
    answer4.insert(END,Accuracy)
    answer4.insert(END,'\n')
    answer4.pack()
    
    answer1.delete(1.0,END)
    for i in range(len(filelist)):
        answer1.insert(END,filelist[i])
        answer1.insert(END,'\n')
    answer1.pack()
    
    answer2.delete(1.0,END)
    for i in range(len(Testing_filelist)):
        answer2.insert(END,Testing_filelist[i])
        answer2.insert(END,'\n')
    answer2.pack()
    
def PrRun(k):
    path=os.getcwd()
    path+="\\Answerfilenew.xlsx"
    WB=load_workbook(path)
    answer3.delete(1.0,END)
    confusionMatrix = [[0 for i in range(6)] for j in range(6)] 

    for i in Testing_filelist:
        sheetname=i.replace("\\", " ")
        ws=WB[sheetname]                
        Answerfrequency={}
        rr=1
        for kkk in range(k):   
            val=ws.cell(row=rr,column=2).value
            f=val.split("\\")
            if f[0] not in Answerfrequency:
                Answerfrequency[f[0]]=1
            else:
                Answerfrequency[f[0]]+=1
            rr+=1
        
        maxfrequency=1
        key=""
        for ky, v in Answerfrequency.items():
            if v>maxfrequency:
                maxfrequency=v
                key=ky
        if key=="":
            val=ws.cell(row=1,column=2).value
            f=val.split("\\")
            key=f[0]
        
        ws.cell(row=2,column=6).value= key
        Answerfrequency.clear()   
        if ws.cell(row=2,column=5).value=="athletics":
            r=0
        elif ws.cell(row=2,column=5).value=="cricket":
            r=1
        elif ws.cell(row=2,column=5).value=="football":
            r=2
        elif ws.cell(row=2,column=5).value=="rugby":
            r=3
        elif ws.cell(row=2,column=5).value=="tennis":
            r=4
            
        if ws.cell(row=2,column=6).value=="athletics":
            c=0
        elif ws.cell(row=2,column=6).value=="cricket":
            c=1
        elif ws.cell(row=2,column=6).value=="football":
            c=2
        elif ws.cell(row=2,column=6).value=="rugby":
            c=3
        elif ws.cell(row=2,column=6).value=="tennis":
            c=4
        answer3.insert(END,ws.cell(row=1,column=1).value)
        answer3.insert(END,'       ---       ')
        answer3.insert(END,ws.cell(row=2,column=6).value)
        answer3.insert(END,'\n')
        
        confusionMatrix[r][c]+=1
        confusionMatrix[5][5]=0
        for x in range(5):
            confusionMatrix[5][x]=0
            confusionMatrix[x][5]=0
            for y in range(5):
                confusionMatrix[5][x]+=confusionMatrix[y][x]
                confusionMatrix[x][5]+=confusionMatrix[x][y]
            
            confusionMatrix[5][5]+=confusionMatrix[5][x]
            
    WB.save("Answerfilenew.xlsx")
    # print(confusionMatrix)
    answer3.pack()
    
    Accuracy=0
    for i in range(len(confusionMatrix)-1):
        Accuracy+=confusionMatrix[i][i]

    Accuracy=Accuracy/(confusionMatrix[5][5])
    # print(Accuracy)
    answer4.delete(1.0,END)
    answer4.insert(END,'Rows = Actual ')
    answer4.insert(END,'\n')
    answer4.insert(END,'Columns = Predicted ')
    answer4.insert(END,'\n')
    answer4.insert(END,'\n')
    answer4.insert(END,'              ')
    answer4.insert(END,"Athletics   ")
    answer4.insert(END,"Cricket     ")
    answer4.insert(END,"Football    ")
    answer4.insert(END,"Rugby     ")
    answer4.insert(END,"Tennis    ")
    answer4.insert(END,"Total     ")
    answer4.insert(END,'\n')
    for i in range(len(confusionMatrix)):
        if i ==0:
            answer4.insert(END,"Athletics    ")
        elif i==1:
            answer4.insert(END,"Cricket       ")
        elif i==2:
            answer4.insert(END,"Football      ")
        elif i==3:
            answer4.insert(END,"Rugby         ")
        elif i==4:
            answer4.insert(END,"Tennis        ")
        elif i==5:
            answer4.insert(END,"Total         ")
        for j in range(len(confusionMatrix)):
            answer4.insert(END,confusionMatrix[i][j])
            if (confusionMatrix[i][j]<10):
                answer4.insert(END,'               ')
            else:
                answer4.insert(END,'              ')
        answer4.insert(END,'\n')
        
    answer4.insert(END,'\n')
    answer4.insert(END,'Accuracy  : ')
    answer4.insert(END,Accuracy)
    answer4.insert(END,'\n')
    answer4.pack()
    

window=Tk()
window.title("K-Nearest Neighbour Classifier")
window.geometry("800x800")
optionsofK=[i+1 for i in range(50) ]
optionsoftrain=[0.0,0.1,0.2,0.3,0.4,0.5,0.6,0.7,0.8,0.9]
optionsoftest=[0.0,0.1,0.2,0.3,0.4,0.5]

oldK=0
oldtr=0.0
oldtes=0.0
def classify():
    global oldK
    global oldtes
    global oldtr
    k=valueofK.get()
    tr=train_ratio.get()
    ts=test_ratio.get()
    if tr+ts!=1.0:
        messagebox.showerror("Split Ratio Error",  "Split ratio (train +test) must be equal to 1.0")
    else:
        if (oldtes!=ts) and (oldtr!=tr):
            programRun(k,tr,ts)
            oldk=k
            oldtr=tr
            oldtes=ts
        elif (oldtes==ts) and (oldtr==tr):
            PrRun(k)
            oldK=k
    

label=Label(window,text="K-Nearest Neighbour Classifier",bg="black",fg="white",padx=5,pady=5)
label.config(font=("Arial",18))
label.pack(fill="x")
h1=Frame(window)

label1=Label(h1,text="Value of K",bg="red",fg="white",padx=10,pady=10)
# label1.config(font=("Arial",18))
label1.grid(row=0,column=0,padx=10,pady=10,sticky="nsew")
label1=Label(h1,text="Training Ratio",bg="red",fg="white",padx=10,pady=10)
# label1.config(font=("Arial",18))
label1.grid(row=0,column=1,padx=10,pady=10,sticky="nsew")
label1=Label(h1,text="Testing Ratio",bg="red",fg="white",padx=10,pady=10)
# label1.config(font=("Arial",18))
label1.grid(row=0,column=2,padx=10,pady=10,sticky="nsew")

h1.grid_columnconfigure(0,weight=1)
h1.grid_columnconfigure(1,weight=1)
h1.grid_columnconfigure(2,weight=1)
h1.pack(fill="x")


h2=Frame(window)

valueofK = IntVar(h2)
valueofK.set(optionsofK[2]) # default value

train_ratio = DoubleVar(h2)
train_ratio.set(optionsoftrain[7]) # default value

test_ratio = DoubleVar(h2)
test_ratio.set(optionsoftest[3]) # default value


w1 = OptionMenu(h2, valueofK, *optionsofK)
w1.grid(row=0,column=0,padx=10,pady=10,sticky="nsew")
w1.configure(bg="black", fg="white",activebackground="white",padx=10,pady=10)

w2 = OptionMenu(h2, train_ratio, *optionsoftrain)
w2.grid(row=0,column=1,padx=10,pady=10,sticky="nsew")
w2.configure(bg="black", fg="white",activebackground="white",padx=10,pady=10)

w3 = OptionMenu(h2, test_ratio, *optionsoftest)
w3.grid(row=0,column=2,padx=10,pady=10,sticky="nsew")
w3.configure(bg="black", fg="white",activebackground="white",padx=10,pady=10)


h2.grid_columnconfigure(0,weight=1)
h2.grid_columnconfigure(1,weight=1)
h2.grid_columnconfigure(2,weight=1)
h2.pack(fill="x")

h3=Frame(window)
button = Button(h3, text="Classify", command=classify)
button.grid(row=0,column=0,padx=5,pady=5,sticky="nsew")
button.configure(bg="black", fg="white",bd=5,activebackground="white")
h3.grid_columnconfigure(0,weight=1)
h3.pack()

frame2=Frame(window)
frame2.pack(fill="both",expand="true")

tablayout=Notebook(frame2)

#tab1
tab1=Frame(tablayout)
tab1.pack(fill="both")

scroll=Scrollbar(tab1)
scroll.pack(side=RIGHT,fill=Y)
answer1=Text(tab1,yscrollcommand=scroll.set,insertborderwidth=3, bd=5,width = 40, font=("Helvetica", 15, "bold"))
scroll.config(command=answer1.yview)
answer1.pack(expand="true", fill="both" )

tablayout.add(tab1,text="Training Filelist")

#tab2
tab2=Frame(tablayout)
tab2.pack(fill="both")


scroll=Scrollbar(tab2)
scroll.pack(side=RIGHT,fill=Y)
answer2=Text(tab2,yscrollcommand=scroll.set,insertborderwidth=3, bd=5,width = 40, font=("Helvetica", 15, "bold"))
scroll.config(command=answer2.yview)
answer2.pack(expand="true", fill="both" )
        
tablayout.add(tab2,text="Testing Filelist")

#tab3
tab3=Frame(tablayout)
tab3.pack(fill="both")

scroll=Scrollbar(tab3)
scroll.pack(side=RIGHT,fill=Y)
answer3=Text(tab3,yscrollcommand=scroll.set,insertborderwidth=3, bd=5,width = 40, font=("Helvetica", 15, "bold"))
scroll.config(command=answer3.yview)
answer3.pack(expand="true", fill="both" )
        
tablayout.add(tab3,text="Actual file - Predicted class")

#tab4
tab4=Frame(tablayout)
tab4.pack(fill="both")

scroll=Scrollbar(tab4)
scroll.pack(side=RIGHT,fill=Y)
answer4=Text(tab4,yscrollcommand=scroll.set,insertborderwidth=3, bd=5,width = 40, font=("Helvetica", 15, "bold"))
scroll.config(command=answer4.yview)
answer4.pack(expand="true", fill="both" )
        
tablayout.add(tab4,text="Confusion Matrix")

tablayout.pack(fill="both")

window.mainloop()

